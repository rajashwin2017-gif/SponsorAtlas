#!/usr/bin/env python3
"""
SponsorAtlas data import script.

Reads:
  - UK Home Office Sponsor Register CSV
  - FOI CoS Activity Excel (2025 data)

Outputs to prisma/data/:
  - sponsors.json      — enriched sponsor records
  - city-stats.json    — pre-computed city aggregates
  - route-stats.json   — pre-computed route aggregates
  - import-summary.json — run stats

Usage:
  python3 scripts/import-data.py \
    --register  "path/to/Worker_and_Temporary_Worker.csv" \
    --foi       "path/to/FOI.xlsx" \
    --out       prisma/data
"""

import argparse
import csv
import hashlib
import json
import os
import re
import sys
from collections import defaultdict
from typing import Optional

# --------------------------------------------------------------------------- #
# Configuration
# --------------------------------------------------------------------------- #

# Canonical city name mapping — normalise variants to a single form
CITY_ALIASES: dict[str, str] = {
    "london,": "London",
    "london, england": "London",
    "london, england,": "London",
    "greater london": "London",
    "city of london": "London",
    "richmond, london": "London",
    "barking, london": "London",
    "hackney, london": "London",
    "croydon, london": "London",
}

# Industry classification — keyword → industry
INDUSTRY_KEYWORDS: list = [
    (["nhs", "hospital", "trust", "health", "medical", "clinic", "care", "dental",
      "pharmacy", "nursing", "therapeutic", "surgery", "wellbeing", "wellness",
      "physio", "optic", "optom", "radiolog", "oncol", "paediatric", "midwif"],
     "Healthcare"),
    (["university", "college", "school", "academy", "education", "learning",
      "institute", "training", "tutor", "lingua", "nursery", "montessori"],
     "Education"),
    (["bank", "financial", "finance", "investment", "capital", "asset", "fund",
      "insurance", "wealth", "credit", "mortgage", "trading", "forex", "hedge",
      "securities", "lloyds", "barclays", "hsbc", "natwest"],
     "Finance"),
    (["technolog", "software", "digital", "cyber", "data", "cloud", "ai ",
      " ai,", "machine learning", "analytics", "devops", "saas", "paas",
      "platform", "systems", "solutions", "computing", "network", "infosys",
      "tata consultancy", "cognizant", "capgemini", "accenture", "wipro"],
     "Technology"),
    (["engineering", "engineer", "manufactur", "industrial", "aerospace",
      "automotive", "mechanical", "electrical", "civil", "structural",
      "process", "chemical", "materials"],
     "Engineering"),
    (["construction", "build", "contractor", "property", "real estate",
      "housing", "architect", "surveyor", "planning", "infrastructure",
      "facilities"],
     "Construction"),
    (["consult", "advisory", "strategy", "management", "deloitte", "kpmg",
      "pwc", "ernst", "mckinsey", "bain", "boston consulting"],
     "Consulting"),
    (["law", "legal", "solicitor", "barrister", "chambers", "attorney"],
     "Legal"),
    (["hotel", "hospitality", "restaurant", "catering", "food", "beverage",
      "pub", "bar ", "cafe", "bakery", "kitchen"],
     "Hospitality"),
    (["retail", "shop", "store", "supermarket", "ecommerce", "e-commerce",
      "marketplace"],
     "Retail"),
    (["research", "science", "bio", "pharma", "lab ", "laboratory",
      "genomic", "clinical", "therapeutics", "astrazeneca", "glaxo", "pfizer"],
     "Life Sciences"),
    (["logistics", "transport", "shipping", "freight", "courier", "warehouse",
      "supply chain", "distribution"],
     "Logistics"),
    (["creative", "media", "publish", "broadcast", "film", "tv ", "television",
      "music", "art ", "design", "advertis", "marketing", "pr ", " pr,"],
     "Creative & Media"),
    (["charity", "non-profit", "ngo", "foundation", "trust ", "voluntary",
      "social care", "community"],
     "Charity & Non-Profit"),
    (["religious", "church", "mosque", "temple", "faith", "minister",
      "worship", "diocese", "parish"],
     "Religious"),
]

DEFAULT_INDUSTRY = "Other"


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #

def make_id(name: str) -> str:
    """Deterministic UUID-shaped ID from org name."""
    h = hashlib.md5(name.upper().encode()).hexdigest()
    return f"{h[:8]}-{h[8:12]}-{h[12:16]}-{h[16:20]}-{h[20:32]}"


def normalise_city(raw: str) -> str:
    raw = raw.strip()
    lower = raw.lower()
    if lower in CITY_ALIASES:
        return CITY_ALIASES[lower]
    # Capitalise properly
    return raw.title() if raw else ""


def classify_industry(name: str, routes: list[str]) -> str:
    name_lower = name.lower()
    route_lower = " ".join(routes).lower()
    combined = name_lower + " " + route_lower

    # Health & Care route is a strong signal
    if "health" in route_lower and "care" in route_lower:
        return "Healthcare"

    for keywords, industry in INDUSTRY_KEYWORDS:
        for kw in keywords:
            if kw in combined:
                return industry

    return DEFAULT_INDUSTRY


def parse_rating(raw: str) -> tuple[str, str]:
    """Returns (rating, ratingType)."""
    raw = raw.strip()
    rating = "B" if "B rating" in raw else "A"
    if raw.startswith("Temporary Worker"):
        return rating, "Temporary Worker"
    if "Provisional" in raw:
        return "Provisional", "Worker"
    return rating, "Worker"


def compute_strength_score(rating: str, cos_sw: Optional[int], cos_sw_supp: bool,
                            cos_gbm: Optional[int], cos_gbm_supp: bool,
                            route_count: int) -> int:
    # Rating: 30pts
    rating_pts = {"A": 30, "B": 10, "Provisional": 15}.get(rating, 10)

    # CoS volume: 40pts (log scale)
    total = (cos_sw or 0) + (cos_gbm or 0)
    supp = cos_sw_supp or cos_gbm_supp
    if total >= 500:
        cos_pts = 40
    elif total >= 200:
        cos_pts = 34
    elif total >= 100:
        cos_pts = 28
    elif total >= 50:
        cos_pts = 22
    elif total >= 20:
        cos_pts = 16
    elif total >= 10:
        cos_pts = 10
    elif total > 0:
        cos_pts = 5
    elif supp:
        cos_pts = 4  # suppressed = confirmed active but small
    else:
        cos_pts = 0

    # Route breadth: 15pts
    route_pts = 15 if route_count >= 3 else (10 if route_count == 2 else 5)

    # GBM activity: 15pts
    gbm_pts = 15 if (cos_gbm and cos_gbm > 0) or cos_gbm_supp else 0

    return min(100, rating_pts + cos_pts + route_pts + gbm_pts)


def compute_opportunity_score(cos_sw: Optional[int], cos_sw_supp: bool,
                               cos_gbm: Optional[int], cos_gbm_supp: bool,
                               city_normalised: str, route_count: int,
                               tier: str) -> int:
    total = (cos_sw or 0) + (cos_gbm or 0)
    supp = cos_sw_supp or cos_gbm_supp

    # CoS volume: 50pts
    if total >= 500:
        cos_pts = 50
    elif total >= 200:
        cos_pts = 42
    elif total >= 100:
        cos_pts = 34
    elif total >= 50:
        cos_pts = 26
    elif total >= 20:
        cos_pts = 18
    elif total >= 10:
        cos_pts = 10
    elif total > 0:
        cos_pts = 5
    elif supp:
        cos_pts = 4
    else:
        cos_pts = 0

    # City demand: 20pts
    TOP_CITIES = {"london": 20, "manchester": 15, "birmingham": 15,
                  "edinburgh": 14, "glasgow": 14, "bristol": 13,
                  "cambridge": 14, "oxford": 14, "leeds": 13,
                  "reading": 12, "coventry": 11, "nottingham": 11,
                  "leicester": 11, "sheffield": 10, "liverpool": 10}
    city_pts = TOP_CITIES.get(city_normalised.lower(), 5)

    # Route breadth: 15pts
    route_pts = 15 if route_count >= 3 else (10 if route_count == 2 else 5)

    # Tier: 15pts
    tier_pts = {"Platinum": 15, "Gold": 12, "Silver": 8, "Bronze": 4}.get(tier, 2)

    return min(100, cos_pts + city_pts + route_pts + tier_pts)


def assign_tier(cos_total: int) -> str:
    if cos_total >= 500:
        return "Platinum"
    if cos_total >= 200:
        return "Gold"
    if cos_total >= 50:
        return "Silver"
    if cos_total >= 10:
        return "Bronze"
    return "Active"


def assign_hiring_activity(cos_total: int, cos_suppressed: bool) -> str:
    if cos_total >= 200:
        return "Very High"
    if cos_total >= 50:
        return "High"
    if cos_total >= 10:
        return "Medium"
    if cos_total > 0 or cos_suppressed:
        return "Low"
    return "Inactive"


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--register", required=True, help="Path to sponsor register CSV")
    parser.add_argument("--foi", required=True, help="Path to FOI Excel file")
    parser.add_argument("--out", default="prisma/data", help="Output directory")
    args = parser.parse_args()

    try:
        import openpyxl
    except ImportError:
        print("Installing openpyxl...")
        os.system(f"{sys.executable} -m pip install openpyxl -q")
        import openpyxl

    os.makedirs(args.out, exist_ok=True)

    # ── 1. Load FOI data ──────────────────────────────────────────────────── #
    print("Loading FOI data...")
    wb = openpyxl.load_workbook(args.foi)

    sw_cos: dict = {}
    ws_sw = wb["FOI Skilled Worker"]
    for r in range(25, ws_sw.max_row + 1):
        org = ws_sw.cell(r, 3).value
        cos = ws_sw.cell(r, 5).value
        if org and (cos is not None):
            sw_cos[org.strip().upper()] = cos

    gbm_cos: dict = {}
    ws_gbm = wb["FOI Global Business Mobility"]
    for r in range(25, ws_gbm.max_row + 1):
        org = ws_gbm.cell(r, 3).value
        cos = ws_gbm.cell(r, 5).value
        if org and (cos is not None):
            gbm_cos[org.strip().upper()] = cos

    print(f"  Skilled Worker FOI records: {len(sw_cos):,}")
    print(f"  GBM FOI records:            {len(gbm_cos):,}")

    # ── 2. Load and aggregate register ───────────────────────────────────── #
    print("Loading sponsor register...")
    org_data: dict = {}

    with open(args.register, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            name = row["Organisation Name"].strip()
            key = name.upper()
            route = row["Route"].strip()
            city_raw = row["Town/City"].strip()
            county_raw = row["County"].strip()
            rating_raw = row["Type & Rating"].strip()

            if key not in org_data:
                org_data[key] = {
                    "name": name,
                    "key": key,
                    "routes": [],
                    "city": normalise_city(city_raw),
                    "county": county_raw if county_raw and county_raw.upper() != "NULL" else "",
                    "rating_raw": rating_raw,
                }

            if route and route not in org_data[key]["routes"]:
                org_data[key]["routes"].append(route)

    print(f"  Unique sponsors in register: {len(org_data):,}")

    # ── 3. Build enriched records ─────────────────────────────────────────── #
    print("Enriching sponsors...")
    sponsors: list = []

    matched = 0
    for key, d in org_data.items():
        name = d["name"]
        routes = d["routes"]
        city = d["city"]
        county = d["county"]
        rating_raw = d["rating_raw"]
        rating, rating_type = parse_rating(rating_raw)

        # FOI data
        raw_sw = sw_cos.get(key)
        raw_gbm = gbm_cos.get(key)

        cos_sw: Optional[int] = raw_sw if isinstance(raw_sw, (int, float)) else None
        cos_sw_supp: bool = raw_sw == "*"
        cos_gbm: Optional[int] = raw_gbm if isinstance(raw_gbm, (int, float)) else None
        cos_gbm_supp: bool = raw_gbm == "*"

        if raw_sw is not None or raw_gbm is not None:
            matched += 1

        cos_total = (cos_sw or 0) + (cos_gbm or 0)
        any_suppressed = cos_sw_supp or cos_gbm_supp

        tier = assign_tier(cos_total) if cos_total > 0 else ("Active" if any_suppressed else "Inactive")
        hiring_activity = assign_hiring_activity(cos_total, any_suppressed)

        strength_score = compute_strength_score(
            rating, cos_sw, cos_sw_supp, cos_gbm, cos_gbm_supp, len(routes)
        )
        opportunity_score = compute_opportunity_score(
            cos_sw, cos_sw_supp, cos_gbm, cos_gbm_supp, city, len(routes), tier
        )
        industry = classify_industry(name, routes)

        sponsors.append({
            "id": make_id(name),
            "organisationName": name,
            "town": city,
            "county": county,
            "routes": routes,
            "rating": rating,
            "ratingType": rating_type,
            "licenceStatus": "Active",
            "industryCategory": industry,
            "sponsorTier": tier,
            "hiringActivity": hiring_activity,
            "sponsorStrengthScore": strength_score,
            "opportunityScore": opportunity_score,
            "cos2025Sw": cos_sw,
            "cos2025Gbm": cos_gbm,
            "cos2025SwSuppressed": cos_sw_supp,
            "cos2025GbmSuppressed": cos_gbm_supp,
            "cos2025Total": cos_total if cos_total > 0 else None,
            # Legacy compat fields (used by existing UI)
            "route": routes[0] if routes else "Skilled Worker",
            "hiringLikelihoodScore": strength_score,
            "cosActivity2025": cos_sw or 0,
            "companySize": (
                "Enterprise" if cos_total >= 200
                else "SME" if cos_total >= 20
                else "Startup"
            ),
            "liveJobsCount": 0,
            "lastJobPostedAt": None,
            "suggestedSocCodes": [],
            "sicCode": "",
            "companiesHouseNumber": "",
            "addedDate": "2024-01-01T00:00:00Z",
            "lastUpdated": "2025-06-16T00:00:00Z",
        })

    print(f"  FOI matched: {matched:,} / {len(org_data):,} ({matched/len(org_data)*100:.1f}%)")

    # Sort by opportunity score desc for default ordering
    sponsors.sort(key=lambda s: -(s["opportunityScore"]))

    # ── 4. City stats ─────────────────────────────────────────────────────── #
    print("Computing city stats...")
    city_map: dict = defaultdict(lambda: {
        "totalSponsors": 0, "cos2025Total": 0,
        "routes": defaultdict(int), "industries": defaultdict(int),
        "tiers": defaultdict(int),
    })

    for s in sponsors:
        city = s["town"] or "Unknown"
        c = city_map[city]
        c["totalSponsors"] += 1
        c["cos2025Total"] += s["cos2025Total"] or 0
        for r in s["routes"]:
            c["routes"][r] += 1
        c["industries"][s["industryCategory"]] += 1
        c["tiers"][s["sponsorTier"]] += 1

    city_stats = []
    for city, c in city_map.items():
        if city and c["totalSponsors"] >= 5:  # filter tiny villages
            top_industry = max(c["industries"], key=c["industries"].get) if c["industries"] else None
            top_route = max(c["routes"], key=c["routes"].get) if c["routes"] else None
            city_stats.append({
                "cityName": city,
                "normalised": city.lower(),
                "totalSponsors": c["totalSponsors"],
                "cos2025Total": c["cos2025Total"],
                "topIndustry": top_industry,
                "topRoute": top_route,
                "tierBreakdown": dict(c["tiers"]),
            })

    city_stats.sort(key=lambda c: -c["cos2025Total"])
    print(f"  Cities with 5+ sponsors: {len(city_stats):,}")

    # ── 5. Route stats ────────────────────────────────────────────────────── #
    print("Computing route stats...")
    route_map: dict = defaultdict(lambda: {
        "totalSponsors": 0, "cos2025Total": 0, "topCities": defaultdict(int),
    })
    for s in sponsors:
        for r in s["routes"]:
            rm = route_map[r]
            rm["totalSponsors"] += 1
            rm["cos2025Total"] += s["cos2025Total"] or 0
            rm["topCities"][s["town"]] += 1

    route_stats = []
    for route, rm in route_map.items():
        top_cities = sorted(rm["topCities"].items(), key=lambda x: -x[1])[:5]
        route_stats.append({
            "route": route,
            "slug": re.sub(r"[^a-z0-9]+", "-", route.lower()).strip("-"),
            "totalSponsors": rm["totalSponsors"],
            "cos2025Total": rm["cos2025Total"],
            "topCities": [{"city": c, "count": n} for c, n in top_cities],
        })
    route_stats.sort(key=lambda r: -r["totalSponsors"])

    # ── 6. Summary ────────────────────────────────────────────────────────── #
    tier_counts = defaultdict(int)
    activity_counts = defaultdict(int)
    industry_counts = defaultdict(int)
    for s in sponsors:
        tier_counts[s["sponsorTier"]] += 1
        activity_counts[s["hiringActivity"]] += 1
        industry_counts[s["industryCategory"]] += 1

    summary = {
        "totalSponsors": len(sponsors),
        "matchedToFOI": matched,
        "matchRate": round(matched / len(sponsors) * 100, 1),
        "totalCos2025Sw": sum(s["cos2025Sw"] or 0 for s in sponsors),
        "totalCos2025Gbm": sum(s["cos2025Gbm"] or 0 for s in sponsors),
        "tierBreakdown": dict(tier_counts),
        "activityBreakdown": dict(activity_counts),
        "topIndustries": sorted(industry_counts.items(), key=lambda x: -x[1])[:10],
        "generatedAt": "2025-06-22T00:00:00Z",
        "dataSource": {
            "register": "Home Office Sponsor Register (June 2025)",
            "foi": "FOI 2026/03173 — 2025 CoS data (extracted March 2026)",
        },
    }

    # ── 7. Write output ───────────────────────────────────────────────────── #
    print("Writing output files...")

    out_sponsors = os.path.join(args.out, "sponsors.json")
    with open(out_sponsors, "w", encoding="utf-8") as f:
        json.dump(sponsors, f, separators=(",", ":"))
    size_mb = os.path.getsize(out_sponsors) / 1_048_576
    print(f"  sponsors.json:     {len(sponsors):,} records, {size_mb:.1f} MB")

    out_cities = os.path.join(args.out, "city-stats.json")
    with open(out_cities, "w", encoding="utf-8") as f:
        json.dump(city_stats, f, separators=(",", ":"))
    print(f"  city-stats.json:   {len(city_stats):,} cities")

    out_routes = os.path.join(args.out, "route-stats.json")
    with open(out_routes, "w", encoding="utf-8") as f:
        json.dump(route_stats, f, separators=(",", ":"))
    print(f"  route-stats.json:  {len(route_stats):,} routes")

    out_summary = os.path.join(args.out, "import-summary.json")
    with open(out_summary, "w", encoding="utf-8") as f:
        json.dump(summary, f, indent=2)
    print(f"  import-summary.json written")

    print("\nDone! Summary:")
    print(f"  {summary['totalSponsors']:,} sponsors")
    print(f"  {summary['matchedToFOI']:,} matched to FOI ({summary['matchRate']}%)")
    for tier, count in sorted(tier_counts.items(), key=lambda x: -x[1]):
        print(f"  {tier}: {count:,}")


if __name__ == "__main__":
    main()
